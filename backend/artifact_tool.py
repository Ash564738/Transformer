# artifact_tool.py
from __future__ import annotations
from copy import copy
from pathlib import Path
import openpyxl
from openpyxl import Workbook as OpenpyxlWorkbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

class _FormatProxy:
    def __init__(self, cell_range): self._range = cell_range
    def __call__(self, value): self._apply_dict(value)
    def _apply(self, key, value):
        for row in self._range._cells():
            for cell in row: setattr(cell, key, value)
    def _apply_dict(self, options):
        if not isinstance(options, dict): return
        fill = options.get("fill"); font = options.get("font")
        horizontal = options.get("horizontal_alignment"); vertical = options.get("vertical_alignment")
        wrap = options.get("wrap_text"); row_height = options.get("row_height")
        number_format = options.get("number_format")
        for row in self._range._cells():
            for cell in row:
                if fill: cell.fill = PatternFill(fill_type="solid", fgColor=str(fill).replace("#", ""))
                if isinstance(font, dict):
                    current_font = copy(cell.font)
                    if "bold" in font: current_font.bold = bool(font["bold"])
                    if "italic" in font: current_font.italic = bool(font["italic"])
                    if "size" in font: current_font.sz = float(font["size"])
                    if "color" in font: current_font.color = str(font["color"]).replace("#", "")
                    cell.font = current_font
                alignment = cell.alignment
                cell.alignment = Alignment(
                    horizontal=horizontal if horizontal is not None else alignment.horizontal,
                    vertical=vertical if vertical is not None else alignment.vertical,
                    text_rotation=alignment.text_rotation,
                    wrap_text=bool(wrap) if wrap is not None else alignment.wrap_text,
                    shrink_to_fit=alignment.shrink_to_fit,
                    indent=alignment.indent)
                if number_format is not None: cell.number_format = str(number_format)
        if row_height is not None:
            for row in range(self._range._min_row(), self._range._max_row() + 1):
                self._range._ws.row_dimensions[row].height = row_height
    def apply_config(self, options): self._apply_dict(options)
    @property
    def wrap_text(self): return None
    @wrap_text.setter
    def wrap_text(self, value):
        for row in self._range._cells():
            for cell in row:
                alignment = cell.alignment
                cell.alignment = Alignment(horizontal=alignment.horizontal, vertical=alignment.vertical,
                    text_rotation=alignment.text_rotation, wrap_text=bool(value),
                    shrink_to_fit=alignment.shrink_to_fit, indent=alignment.indent)
    @property
    def column_width(self): return None
    @column_width.setter
    def column_width(self, value):
        min_col = self._range._min_col(); max_col = self._range._max_col()
        for col in range(min_col, max_col + 1): self._range._ws.column_dimensions[get_column_letter(col)].width = float(value)
    @property
    def number_format(self):
        for row in self._range._cells():
            for cell in row: return cell.number_format
        return None
    @number_format.setter
    def number_format(self, value):
        for row in self._range._cells():
            for cell in row: cell.number_format = value
    def autofit_columns(self):
        ws = self._range._ws
        for col in range(self._range._min_col(), self._range._max_col() + 1):
            letter = get_column_letter(col); width = 0
            for cell in ws[letter]:
                value = cell.value
                if value is None: continue
                width = max(width, len(str(value)))
            ws.column_dimensions[letter].width = min(max(width + 2, 10), 60)
    def __setattr__(self, name, value):
        if name == "_range" or name.startswith("_"): object.__setattr__(self, name, value); return
        if name == "wrap_text": type(self).wrap_text.fset(self, value); return
        if name == "column_width": type(self).column_width.fset(self, value); return
        if name == "number_format": type(self).number_format.fset(self, value); return
        object.__setattr__(self, name, value)

class _Range:
    def __init__(self, worksheet, range_string):
        object.__setattr__(self, "_ws", worksheet)
        object.__setattr__(self, "range_string", str(range_string))
        object.__setattr__(self, "_format_proxy", _FormatProxy(self))
    @property
    def ref(self): return self.range_string
    @property
    def format(self): return self._format_proxy
    @format.setter
    def format(self, value): self._format_proxy.apply_config(value)
    def _range(self): return self._ws[self.ref]
    def _cells(self):
        cells = self._ws[self.ref]
        if isinstance(cells, tuple):
            if cells and isinstance(cells[0], tuple): return cells
            return (cells,)
        return ((cells,),)
    def _bounds(self):
        min_col, min_row, max_col, max_row = range_boundaries(self.ref)
        return min_col, min_row, max_col, max_row
    def _min_col(self): return self._bounds()[0]
    def _max_col(self): return self._bounds()[2]
    def _min_row(self): return self._bounds()[1]
    def _max_row(self): return self._bounds()[3]
    @property
    def values(self): return [[c.value for c in row] for row in self._cells()]
    @values.setter
    def values(self, data):
        if data is None: return
        rows = list(data)
        for r_off, row in enumerate(rows):
            for c_off, value in enumerate(row):
                self._ws.cell(row=self._min_row() + r_off, column=self._min_col() + c_off, value=value)
    def merge(self): self._ws.merge_cells(self.ref)
    def unmerge(self): self._ws.unmerge_cells(self.ref)
    def clear(self):
        for row in self._cells():
            for cell in row: cell.value = None
    def write(self, rows):
        rows = list(rows)
        if not rows: return
        cols = max(len(r) for r in rows)
        normalized = [list(r) + [None] * (cols - len(r)) for r in rows]
        for r_off, row in enumerate(normalized):
            for c_off, value in enumerate(row):
                self._ws.cell(row=self._min_row() + r_off, column=self._min_col() + c_off, value=value)
    def set_number_format(self, value): self.format.number_format = value
    def autofit_columns(self): self.format.autofit_columns()

class _FreezePanes:
    def __init__(self, worksheet): self._ws = worksheet
    def freeze_rows(self, count): self._ws.freeze_panes = f"A{int(count) + 1}"
    def freeze_columns(self, count): self._ws.freeze_panes = f"{get_column_letter(int(count) + 1)}1"
    def unfreeze(self): self._ws.freeze_panes = None

class _ChartLegend:
    def __init__(self, chart): self._chart = chart
    @property
    def position(self): return self._chart.legend.position if self._chart.legend else None
    @position.setter
    def position(self, value):
        if self._chart.legend is None: self._chart.legend = openpyxl.chart.legend.Legend()
        self._chart.legend.position = value

class _ChartProxy:
    def __init__(self, chart): self._chart = chart
    @property
    def title_text(self): return self._chart.title
    @title_text.setter
    def title_text(self, value): self._chart.title = value
    @property
    def has_legend(self): return self._chart.legend is not None
    @has_legend.setter
    def has_legend(self, value):
        if value and self._chart.legend is None: self._chart.legend = openpyxl.chart.legend.Legend()
        elif not value: self._chart.legend = None
    @property
    def legend(self): return _ChartLegend(self._chart)
    def set_position(self, start, end=None):
        self._chart.anchor = start
        if end:
            from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
            col_s, row_s = coordinate_from_string(end)
            start_col_s, start_row = coordinate_from_string(start)
            self._chart.width = max(7.0, (column_index_from_string(col_s) - column_index_from_string(start_col_s)) * 1.1)
            self._chart.height = max(5.0, (int(row_s) - int(start_row)) * 0.35)

class _Charts:
    def __init__(self, sheet):
        self.sheet = sheet
        self._charts = []
    def add(self, chart_type, source_range, title=None):
        ws = self.sheet._ws
        if isinstance(source_range, _Range):
            min_col = source_range._min_col(); max_col = source_range._max_col()
            min_row = source_range._min_row(); max_row = source_range._max_row()
            if chart_type == "line": chart = LineChart()
            elif chart_type == "pie": chart = PieChart()
            else: chart = BarChart()
            chart.title = title or ""
            data = Reference(ws, min_col=min_col, max_col=max_col, min_row=min_row, max_row=max_row)
            cats = Reference(ws, min_col=min_col, min_row=min_row + 1, max_row=max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
        else:
            chart = BarChart() if chart_type not in {"line", "pie"} else LineChart() if chart_type == "line" else PieChart()
            chart.title = title or ""
        ws.add_chart(chart, "J10")
        proxy = _ChartProxy(chart)
        self._charts.append(proxy)
        return proxy
    def delete_all(self):
        self.sheet._ws._charts = []
        self._charts.clear()

class _Sheet:
    def __init__(self, ws):
        self._ws = ws
        self.charts = _Charts(self)
        self.freeze_panes = _FreezePanes(ws)
        self.name = ws.title
    def get_range(self, range_string): return _Range(self._ws, range_string)
    def get_range_by_indexes(self, row_index, column_index, row_count, column_count):
        sr, sc = int(row_index) + 1, int(column_index) + 1
        er, ec = sr + int(row_count) - 1, sc + int(column_count) - 1
        return _Range(self._ws, f"{get_column_letter(sc)}{sr}:{get_column_letter(ec)}{er}")
    def get_cell(self, row, col): return _Range(self._ws, f"{get_column_letter(int(col) + 1)}{int(row) + 1}")
    def merge_cells(self, range_string): self._ws.merge_cells(range_string)

class _Worksheets:
    def __init__(self, workbook):
        self._workbook = workbook
        self.items = []
    def add(self, name):
        if name in self._workbook.sheetnames: del self._workbook[name]
        ws = self._workbook.create_sheet(title=name)
        wrapper = _Sheet(ws)
        self.items.append(wrapper)
        return wrapper
    def get_item(self, name):
        for item in self.items:
            if item._ws.title == name: return item
        if name in self._workbook.sheetnames:
            wrapper = _Sheet(self._workbook[name])
            self.items.append(wrapper)
            return wrapper
        raise KeyError(name)
    def get_or_add(self, name, options=None):
        try: return self.get_item(name)
        except KeyError: return self.add(name)

class Workbook:
    def __init__(self):
        self._wb = OpenpyxlWorkbook()
        default = self._wb.active
        self._wb.remove(default)
        self.worksheets = _Worksheets(self._wb)
    @staticmethod
    def create(): return Workbook()
    def inspect(self, spec):
        kind = spec.get("kind")
        if kind == "sheet":
            data = [{"name": s} for s in self._wb.sheetnames]
        elif kind == "table":
            ref = spec.get("range")
            if "!" in ref: sheet_name, ref = ref.split("!", 1)
            else: sheet_name = self._wb.sheetnames[0]
            ws = self._wb[sheet_name]
            data = [[c.value for c in row] for row in ws[ref]]
        else: data = []
        class Result:
            def __init__(self, obj):
                import json
                self.ndjson = "\n".join(json.dumps(x, ensure_ascii=False, default=str) for x in obj) if isinstance(obj, list) else str(obj)
        return Result(data)
    def render(self, spec): raise NotImplementedError("Rendering is not implemented in this compatibility wrapper.")

class SpreadsheetFile:
    @staticmethod
    def export_xlsx(workbook): return _ExportResult(workbook._wb)
    @staticmethod
    def import_xlsx(blob): raise NotImplementedError("Import is not implemented in this compatibility wrapper.")

class _ExportResult:
    def __init__(self, workbook): self._wb = workbook
    def save(self, path):
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        self._wb.save(path)